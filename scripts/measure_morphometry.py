#!/usr/bin/env python3
"""
measure_morphometry.py  —  digit morphometry measurement on a finished segmentation
=====================================================================================
One command: a segmentation mask + its grayscale volume -> socket / phalanx / bone
morphometry (voxels AND mm) + a 7-class annotated label volume for NiiVue.

This is the measurement counterpart to `segment_microct.py`. It does NOT do any
segmentation: it takes the `<case>.nii.gz` mask and `<case>_0000.nii.gz` grayscale
volume a succeeded segmentation Run already produced, lays them out the way the
vendored `perios/digitpipe_v5` pipeline expects, runs that pipeline as a
subprocess, and consolidates everything it emitted into one flat JSON.

CPU-only. No torch, no GPU.

-------------------------------------------------------------------------------------
QUICK START
-------------------------------------------------------------------------------------
    python scripts/measure_morphometry.py \
      --mask  "<results>/R2__<model>__run7/R2.nii.gz" \
      --image "<results>/R2__<model>__run7/R2_0000.nii.gz" \
      --case  R2 \
      --out   "<results>/R2__morph__m1" \
      --pipeline digitpipe_v5 --skip-viz --spacing-um 4.000059

Outputs in --out:
    labels/<case>.nii.gz            copy of the mask   (pipeline input)
    images/<case>.nii.gz            copy of the volume (pipeline input)
    metrics/*.json                  the pipeline's per-stage metrics
    labelsBoneLength_v2/            the 7-class annotated volume (full resolution)
    output.xlsx                     the pipeline's spreadsheet
    <case>_measurement.json         >>> the consolidated flat result this app reads

-------------------------------------------------------------------------------------
COMMON OPTIONS
-------------------------------------------------------------------------------------
  --pipeline digitpipe_v5   which vendored pipeline under scripts/perios/ to run
  --skip-viz                skip the GIF visualization stage (much faster)
  --spacing-um 4.0          voxel size in MICROMETRES. Used for the voxels->mm
                            conversions. Read "Image Pixel Size (um)" from the
                            *_rec.log (R2/R4 are 4.000059). NOTE: the vendored
                            pipeline hard-codes 4.0 um for ITS OWN output.xlsx;
                            this script recomputes every mm value from --spacing-um
                            so the consolidated JSON is correct even when they differ.

=====================================================================================
METRIC KEY PROVENANCE  (read this before trusting / changing the parser)
=====================================================================================
The key names below were read out of the vendored source at commit
88bedbaa (see scripts/perios/PROVENANCE.md) — specifically
`digitpipe_v5/run_pipeline.py` and `digitpipe_v5/utils.py` — and then CONFIRMED by
executing the real pipeline end to end on a synthetic two-prong phantom (all 8
stages OK; socket detected; bone length and the 7-class annotated volume produced).
The unit scaling below was cross-checked against the pipeline's own output.xlsx:
our Socket_Volume_voxels / Socket_Radius_voxels / Bone_Length_voxels match its
columns exactly, and only the *_mm columns differ — deliberately, because ours use
--spacing-um while the pipeline hard-codes 4.0 um.

What that phantom could NOT confirm: it produced identical 50th/95th-percentile
shrink volumes, so `phalanx_volume` (95th) vs `phalanx_shrunk50_volume` (50th)
were not distinguished by it; on real specimens they will differ.

The parser stays deliberately defensive regardless: every lookup tolerates a
missing file / missing key, and anything it does not recognize is preserved
verbatim under the top-level "raw" key. A degenerate sample (no socket detected —
stage 4 can legitimately find nothing) yields zero/None metrics and NO annotated
volume, but is still a SUCCESS, not an error: the pipeline ran correctly and
"this specimen has no measurable socket" is a real result, not a failure.

Files the pipeline writes into <out>/metrics/ (run_pipeline.py, one per stage):

  downsample_info.json          {"downsample_factor": 2, "samples": [{sample_name,
                                 original_shape, downsampled_shape, original_affine}]}
  hull_metrics_ds.json          [utils.compute_hull_metrics] original_volume,
                                 hull_volume, containment_voxels, containment_ratio,
                                 expansion_ratio, sample_name
  shrink_metrics_ds.json        [utils.compute_shrink_metrics] original_volume,
                                 hull_volume, shrunk_volume, containment_ratio,
                                 hull_expansion, shrunk_expansion, reduction_ratio,
                                 voxels_removed, sample_name, percentile (50 and 95)
  axis_info_ds.json             [utils.compute_major_axis] center, direction, length,
                                 endpoints, explained_variance, sample_name
  socket_metrics_ds.json        [utils.compute_socket_metrics] shrunk_volume,
                                 hull_volume, socket_volume, num_components,
                                 centroid, equivalent_radius, sample_name
  bone_length_metrics_v2_ds.json  [run_pipeline.measure_bone_length_v2] socket_com,
                                 furthest_point, first_intersection,
                                 euclidean_distance_total, bone_length_euclidean,
                                 socket_length_euclidean, socket_volume, bone_volume,
                                 TV_primary, BV_primary, BV_TV_primary,
                                 TV_secondary, BV_secondary, BV_TV_secondary
  bone_length_metrics_v2.json   same keys, UPSCALED to full resolution by stage 6

UNITS — the single most important thing to get right here:
  * Stages 1-5 all run on the DOWNSAMPLED grid (DOWNSAMPLE_FACTOR = 2), so every
    "_ds" metrics file is in DOWNSAMPLED voxels. To reach full-resolution voxels:
      volumes  x factor**3     lengths/radii x factor     coordinates x factor
    (this is exactly what run_pipeline.scale_metrics does, and what stage_7_excel
    re-does inline for socket volume/radius).
  * `bone_length_metrics_v2.json` (no "_ds") is ALREADY full resolution — stage 6
    applied scale_metrics to it. We prefer that file and only fall back to the
    "_ds" one (applying the scaling ourselves) if it is absent.
  * mm = full-resolution voxels x spacing_mm ; mm3 = voxels x spacing_mm**3.

SEMANTIC MAPPINGS that are inferences, not literal key names:
  * `phalanx_volume` <- socket_metrics_ds."shrunk_volume". Inferred from
    07_export_excel.ipynb, which labels that exact field "Phalanx_Volume_voxels".
    (It is the 95th-percentile shrink-wrapped mask.) For cross-checking we ALSO
    emit `phalanx_shrunk50_volume_*` from bone_length "TV_secondary", which is the
    50th-percentile shrink-wrapped mask that stage 5 actually measures BV/TV on.
  * `bone_length` <- "bone_length_euclidean" (v2). The v1 metrics that
    07_export_excel.ipynb also knows about are NOT produced by run_pipeline.py.
  * `euclidean_distance` <- "euclidean_distance_total": the straight-line distance
    from the socket centre of mass to the furthest bone voxel.
  * LINE INSIDE / OUTSIDE: stage 5 walks a Bresenham line from the socket centre of
    mass to that furthest bone voxel and splits it at the FIRST bone voxel it hits.
      - the part after the hit  = "bone_length_euclidean"   -> line_inside_bone_*
      - the part before the hit = "socket_length_euclidean" -> line_outside_bone_*
    So line_inside_bone_* is by definition identical to bone_length_*; both names
    are emitted because the two readings ("how long is the bone" / "how much of the
    line lies inside bone") are asked for separately.

The 7-class annotated volume (<out>/labelsBoneLength_v2/<case>_bone_length_v2.nii.gz)
uses these labels, from run_pipeline.measure_bone_length_v2:
    1 bone mask   2 socket   3 line segment outside bone   4 line segment inside bone
    5 furthest bone point    6 socket centre of mass       7 first intersection
"""
import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import time

# Exact strings digitpipe_v5's run_pipeline.py prints when a stage did not do its
# job. It catches every per-stage exception and still exits 0, so these lines are
# the only honest signal available. Verified against the vendored commit — see
# scripts/perios/PROVENANCE.md; re-check them after re-vendoring.
#   "[FAIL]"                     summary table (~line 704: "  [{status}] {name}: ...")
#   "ERROR in "                  printed at the catch site, carries the exception text
#   "SKIPPED - missing files"    stage 5 could not find a sample's inputs
#   "Socket metrics not found"   stage 5 had no stage-4 output to work from
# NB: "--skip-viz" does NOT produce any of these — it is only an argparse flag —
# so a normal skip-viz run never trips this.
PIPELINE_FAILURE_MARKERS = (
    "[FAIL]",
    "ERROR in ",
    "SKIPPED - missing files",
    "Socket metrics not found",
)

# Label meanings of the annotated volume the pipeline emits (see docstring).
ANNOTATED_LABELS = {
    1: "bone", 2: "socket", 3: "line_outside_bone", 4: "line_inside_bone",
    5: "furthest_point", 6: "socket_centroid", 7: "first_intersection",
}

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PERIOS_ROOT = os.path.join(REPO_ROOT, "scripts", "perios")


def log(*a):
    print(*a, flush=True)


def die(msg, code=1):
    print(f"ERROR: {msg}", file=sys.stderr, flush=True)
    log(f"[error] {msg}")
    sys.exit(code)


# --------------------------------------------------------------------------- setup
def resolve_pipeline(name):
    """Locate the vendored pipeline directory, refusing anything outside perios/."""
    if not re.fullmatch(r"[A-Za-z0-9_.-]+", name or ""):
        die(f"invalid --pipeline name: {name!r}")
    folder = os.path.abspath(os.path.join(PERIOS_ROOT, name))
    if os.path.commonpath([folder, os.path.abspath(PERIOS_ROOT)]) != os.path.abspath(PERIOS_ROOT):
        die(f"--pipeline must name a directory inside {PERIOS_ROOT}")
    script = os.path.join(folder, "run_pipeline.py")
    if not os.path.isfile(script):
        die(f"vendored pipeline not found: {script}. "
            f"Re-vendor it (see scripts/perios/PROVENANCE.md).")
    return folder, script


def stage_inputs(mask, image, out, case):
    """Lay the mask + volume out as <out>/labels/<case>.nii.gz and images/<case>.nii.gz.

    Copies (never moves) — the run's outputs are an immutable provenance record.
    The pipeline matches labels to images purely by filename stem, so both copies
    MUST be named <case>.nii.gz (the grayscale keeps its _0000 name in the run dir).
    """
    for src, kind in ((mask, "mask"), (image, "image")):
        if not os.path.isfile(src):
            die(f"{kind} not found: {src}")
    labels_dir = os.path.join(out, "labels")
    images_dir = os.path.join(out, "images")
    os.makedirs(labels_dir, exist_ok=True)
    os.makedirs(images_dir, exist_ok=True)
    dst_mask = os.path.join(labels_dir, f"{case}.nii.gz")
    dst_image = os.path.join(images_dir, f"{case}.nii.gz")
    shutil.copyfile(mask, dst_mask)
    log(f"[prepare] labels/{case}.nii.gz  <- {mask}")
    shutil.copyfile(image, dst_image)
    log(f"[prepare] images/{case}.nii.gz  <- {image}")
    return dst_mask, dst_image


# ----------------------------------------------------------------- metric parsing
def _load_json(path):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return None


def _pick(records, case):
    """The record for `case` in a list-of-dicts metrics file.

    Matches sample_name exactly, then with a trailing '.nii' (the pipeline is
    inconsistent about stripping it), then case-insensitively. If the file holds
    exactly one record and nothing matched, that record is it — we always feed the
    pipeline a single sample, so a name mismatch is a naming quirk, not ambiguity.
    """
    if not isinstance(records, list) or not records:
        return None
    wanted = {case, f"{case}.nii", f"{case}.nii.gz"}
    for r in records:
        if isinstance(r, dict) and r.get("sample_name") in wanted:
            return r
    low = {w.lower() for w in wanted}
    for r in records:
        if isinstance(r, dict) and str(r.get("sample_name", "")).lower() in low:
            return r
    return records[0] if len(records) == 1 and isinstance(records[0], dict) else None


def _all_for(records, case):
    """Every record for `case` (shrink metrics emit one row per percentile)."""
    if not isinstance(records, list):
        return []
    wanted = {case, f"{case}.nii", f"{case}.nii.gz"}
    hits = [r for r in records
            if isinstance(r, dict) and r.get("sample_name") in wanted]
    return hits or [r for r in records if isinstance(r, dict)]


def _num(v):
    """Coerce to a finite float, else None (JSON can't carry NaN/Inf)."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f if f == f and f not in (float("inf"), float("-inf")) else None


def _scaled(v, factor, power):
    """Full-resolution value from a downsampled one (power 1=length, 3=volume)."""
    n = _num(v)
    return None if n is None else n * (factor ** power)


def _coords(v, factor=1.0):
    """A 3-vector of full-resolution voxel coordinates, or None."""
    if not isinstance(v, (list, tuple)) or len(v) != 3:
        return None
    out = [_num(c) for c in v]
    return None if any(c is None for c in out) else [c * factor for c in out]


def read_xlsx(path, case):
    """The pipeline's own All_Measurements row for `case`, as a plain dict.

    Best-effort: openpyxl first (it is the engine stage 7 writes with, so it is
    guaranteed present whenever the file exists), then pandas, then give up. The
    xlsx is a cross-check on our own arithmetic, never the source of truth — its
    mm columns use the pipeline's hard-coded 4.0 um voxel, ours use --spacing-um.
    """
    if not os.path.isfile(path):
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = "All_Measurements" if "All_Measurements" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return None
        header = [str(h) if h is not None else "" for h in rows[0]]
        for r in rows[1:]:
            row = dict(zip(header, r))
            name = str(row.get("Sample", ""))
            if name in (case, f"{case}.nii", f"{case}.nii.gz") or len(rows) == 2:
                return {k: (v if isinstance(v, (int, float, str, bool, type(None))) else str(v))
                        for k, v in row.items() if k}
        return None
    except Exception:  # noqa: BLE001 — openpyxl missing / unreadable workbook
        pass
    try:
        import pandas as pd
        df = pd.read_excel(path, sheet_name="All_Measurements")
        hit = df[df["Sample"].astype(str).isin([case, f"{case}.nii", f"{case}.nii.gz"])]
        if hit.empty and len(df) == 1:
            hit = df
        if hit.empty:
            return None
        rec = hit.iloc[0].to_dict()
        return {str(k): (None if v != v else v) for k, v in rec.items()}  # NaN -> None
    except Exception:  # noqa: BLE001
        return None


def consolidate(out, case, spacing_um, pipeline_version, skip_viz, timings):
    """Every metric the pipeline emitted, flattened, in both voxels and mm.

    Returns (flat_metrics, raw_records, annotated_nii_path, xlsx_path). Anything
    not explicitly mapped survives under `raw`.
    """
    mdir = os.path.join(out, "metrics")
    spacing_mm = spacing_um / 1000.0
    vox_mm3 = spacing_mm ** 3

    raw = {}
    for stem in ("downsample_info", "hull_metrics_ds", "shrink_metrics_ds",
                 "axis_info_ds", "socket_metrics_ds",
                 "bone_length_metrics_v2_ds", "bone_length_metrics_v2"):
        data = _load_json(os.path.join(mdir, f"{stem}.json"))
        if data is not None:
            raw[stem] = data

    # Downsample factor: the pipeline records it; default to its constant (2).
    ds_info = raw.get("downsample_info") or {}
    factor = _num(ds_info.get("downsample_factor")) or 2.0

    hull = _pick(raw.get("hull_metrics_ds"), case) or {}
    socket = _pick(raw.get("socket_metrics_ds"), case) or {}
    axis = _pick(raw.get("axis_info_ds"), case) or {}
    shrink_rows = _all_for(raw.get("shrink_metrics_ds"), case)

    # bone_length_metrics_v2.json is already full-resolution (stage 6 scaled it);
    # the _ds variant is not. Prefer the scaled one, else scale the _ds one here.
    bone_full = _pick(raw.get("bone_length_metrics_v2"), case)
    bone_ds = _pick(raw.get("bone_length_metrics_v2_ds"), case)
    if bone_full is not None:
        bone, bone_f = bone_full, 1.0
        bone_source = "bone_length_metrics_v2.json (pre-scaled by stage 6)"
    else:
        bone, bone_f = (bone_ds or {}), factor
        bone_source = "bone_length_metrics_v2_ds.json (scaled here)"

    def vol(v, f):   # downsampled voxel count -> full-resolution voxel count
        return _scaled(v, f, 3)

    def length(v, f):
        return _scaled(v, f, 1)

    def mm(v):
        return None if v is None else v * spacing_mm

    def mm3(v):
        return None if v is None else v * vox_mm3

    m = {}

    # ---- socket (socket_metrics_ds.json — downsampled) ----
    m["socket_volume_voxels"] = vol(socket.get("socket_volume"), factor)
    m["socket_volume_mm3"] = mm3(m["socket_volume_voxels"])
    m["socket_radius_voxels"] = length(socket.get("equivalent_radius"), factor)
    m["socket_radius_mm"] = mm(m["socket_radius_voxels"])
    centroid = _coords(socket.get("centroid"), factor)
    m["socket_centroid"] = centroid
    m["socket_centroid_i"] = centroid[0] if centroid else None
    m["socket_centroid_j"] = centroid[1] if centroid else None
    m["socket_centroid_k"] = centroid[2] if centroid else None
    m["socket_num_components"] = _num(socket.get("num_components"))

    # ---- phalanx: the 95th-percentile shrink-wrapped mask (see docstring) ----
    m["phalanx_volume_voxels"] = vol(socket.get("shrunk_volume"), factor)
    m["phalanx_volume_mm3"] = mm3(m["phalanx_volume_voxels"])
    # cross-check: the 50th-percentile mask stage 5 measures BV/TV on
    m["phalanx_shrunk50_volume_voxels"] = vol(bone.get("TV_secondary"), bone_f)
    m["phalanx_shrunk50_volume_mm3"] = mm3(m["phalanx_shrunk50_volume_voxels"])

    # ---- bone length + the line split at the first bone intersection ----
    m["bone_length_voxels"] = length(bone.get("bone_length_euclidean"), bone_f)
    m["bone_length_mm"] = mm(m["bone_length_voxels"])
    m["euclidean_distance_voxels"] = length(bone.get("euclidean_distance_total"), bone_f)
    m["euclidean_distance_mm"] = mm(m["euclidean_distance_voxels"])
    m["line_inside_bone_voxels"] = m["bone_length_voxels"]
    m["line_inside_bone_mm"] = m["bone_length_mm"]
    m["line_outside_bone_voxels"] = length(bone.get("socket_length_euclidean"), bone_f)
    m["line_outside_bone_mm"] = mm(m["line_outside_bone_voxels"])
    m["line_socket_com"] = _coords(bone.get("socket_com"), bone_f)
    m["line_furthest_point"] = _coords(bone.get("furthest_point"), bone_f)
    m["line_first_intersection"] = _coords(bone.get("first_intersection"), bone_f)

    # ---- whole-bone volume + bone-volume-fraction ----
    m["bone_volume_voxels"] = vol(bone.get("bone_volume"), bone_f)
    m["bone_volume_mm3"] = mm3(m["bone_volume_voxels"])
    m["tv_primary_voxels"] = vol(bone.get("TV_primary"), bone_f)
    m["bv_primary_voxels"] = vol(bone.get("BV_primary"), bone_f)
    m["bv_tv_primary"] = _num(bone.get("BV_TV_primary"))
    m["tv_secondary_voxels"] = vol(bone.get("TV_secondary"), bone_f)
    m["bv_secondary_voxels"] = vol(bone.get("BV_secondary"), bone_f)
    m["bv_tv_secondary"] = _num(bone.get("BV_TV_secondary"))

    # ---- hull (hull_metrics_ds.json — downsampled) ----
    m["hull_volume_voxels"] = vol(hull.get("hull_volume"), factor)
    m["hull_volume_mm3"] = mm3(m["hull_volume_voxels"])
    m["hull_original_volume_voxels"] = vol(hull.get("original_volume"), factor)
    m["hull_original_volume_mm3"] = mm3(m["hull_original_volume_voxels"])
    m["hull_containment_voxels"] = vol(hull.get("containment_voxels"), factor)
    m["hull_containment_ratio"] = _num(hull.get("containment_ratio"))
    m["hull_expansion_ratio"] = _num(hull.get("expansion_ratio"))
    # the 3D convex hull from stage 4, which is a different hull to stage 1's
    m["hull3d_volume_voxels"] = vol(socket.get("hull_volume"), factor)
    m["hull3d_volume_mm3"] = mm3(m["hull3d_volume_voxels"])

    # ---- shrink-wrap, one block per percentile ----
    for row in shrink_rows:
        p = row.get("percentile")
        try:
            tag = f"shrink{int(p)}"
        except (TypeError, ValueError):
            continue
        m[f"{tag}_volume_voxels"] = vol(row.get("shrunk_volume"), factor)
        m[f"{tag}_volume_mm3"] = mm3(m[f"{tag}_volume_voxels"])
        m[f"{tag}_containment_ratio"] = _num(row.get("containment_ratio"))
        m[f"{tag}_reduction_ratio"] = _num(row.get("reduction_ratio"))
        m[f"{tag}_hull_expansion"] = _num(row.get("hull_expansion"))
        m[f"{tag}_shrunk_expansion"] = _num(row.get("shrunk_expansion"))
        m[f"{tag}_voxels_removed"] = vol(row.get("voxels_removed"), factor)

    # ---- major axis (axis_info_ds.json — downsampled) ----
    m["major_axis_length_voxels"] = length(axis.get("length"), factor)
    m["major_axis_length_mm"] = mm(m["major_axis_length_voxels"])
    m["major_axis_explained_variance"] = _num(axis.get("explained_variance"))
    m["major_axis_center"] = _coords(axis.get("center"), factor)
    m["major_axis_direction"] = _coords(axis.get("direction"))  # unit vector: no scaling

    # ---- shape / provenance ----
    for s in (ds_info.get("samples") or []):
        if isinstance(s, dict) and s.get("sample_name") in (case, f"{case}.nii"):
            m["volume_shape"] = s.get("original_shape")
            m["downsampled_shape"] = s.get("downsampled_shape")
            break
    m["downsample_factor"] = factor
    m["spacing_um"] = spacing_um
    m["spacing_mm"] = spacing_mm
    m["voxel_volume_mm3"] = vox_mm3
    m["bone_length_metrics_source"] = bone_source

    # ---- the pipeline's own spreadsheet row, for cross-checking ----
    xlsx_path = os.path.join(out, "output.xlsx")
    xlsx_row = read_xlsx(xlsx_path, case)
    if xlsx_row:
        raw["output_xlsx_row"] = xlsx_row
        # Surface its numbers under an xlsx_ prefix so a mismatch is visible in the
        # UI without shadowing our own (correctly-spaced) values.
        for k, v in xlsx_row.items():
            key = "xlsx_" + re.sub(r"[^A-Za-z0-9]+", "_", str(k)).strip("_").lower()
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                continue
            m.setdefault(key, _num(v))

    annotated = os.path.join(out, "labelsBoneLength_v2", f"{case}_bone_length_v2.nii.gz")
    if not os.path.isfile(annotated):
        alt = os.path.join(out, "labelsBoneLength_v2_ds", f"{case}_bone_length_v2.nii.gz")
        annotated = alt if os.path.isfile(alt) else None

    m["pipeline_version"] = pipeline_version
    m["skip_viz"] = bool(skip_viz)
    m.update({f"{k}_seconds": v for k, v in timings.items()})

    # Drop keys that came out entirely empty so the JSON stays readable, but keep
    # the required headline metrics even when null (callers key off their presence).
    required = {
        "socket_volume_voxels", "socket_volume_mm3", "socket_radius_voxels",
        "socket_radius_mm", "socket_centroid", "phalanx_volume_voxels",
        "phalanx_volume_mm3", "bone_length_voxels", "bone_length_mm",
        "euclidean_distance_voxels", "euclidean_distance_mm",
    }
    flat = {k: v for k, v in m.items() if v is not None or k in required}
    return flat, raw, annotated, (xlsx_path if os.path.isfile(xlsx_path) else None)


def capture_env(pipeline_version, timings):
    """Host debrief. Reuses the app's sysinfo when the package is importable
    (the worker spawns us with the same interpreter), else stdlib only."""
    env = {}
    try:
        sys.path.insert(0, os.path.join(REPO_ROOT, "src"))
        from microct_lab.sysinfo import host_info  # type: ignore
        env.update(host_info())
    except Exception:  # noqa: BLE001 — running standalone, outside the app
        import platform
        import socket
        env.update({
            "host": socket.gethostname(),
            "platform": platform.platform(),
            "os": platform.system(),
            "python": platform.python_version(),
            "cpu": platform.processor() or platform.machine(),
            "logical_cores": os.cpu_count(),
        })
    env["device"] = "cpu"  # the morphometry pipeline is CPU-only by construction
    env["pipeline_version"] = pipeline_version
    for mod in ("numpy", "scipy", "skimage", "sklearn", "nibabel", "pandas", "openpyxl"):
        try:
            env[f"{mod}_version"] = getattr(__import__(mod), "__version__", "?")
        except Exception:  # noqa: BLE001 — not installed here; the child may still have it
            pass
    try:
        import psutil
        mi = psutil.Process().memory_info()
        env["peak_ram_mb"] = round(getattr(mi, "peak_wset", getattr(mi, "rss", 0)) / 1e6, 1)
    except Exception:  # noqa: BLE001
        pass
    env.update({f"{k}_seconds": v for k, v in timings.items()})
    return env


def main():
    ap = argparse.ArgumentParser(
        description="segmentation mask -> digit morphometry (perios digitpipe)",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    ap.add_argument("--mask", required=True, help="Binary mask .nii.gz from a succeeded run")
    ap.add_argument("--image", required=True, help="Grayscale volume .nii.gz (_0000) from that run")
    ap.add_argument("--case", required=True, help="Case name (prefixes the outputs)")
    ap.add_argument("--out", required=True, help="Output directory")
    ap.add_argument("--pipeline", default="digitpipe_v5",
                    help="Vendored pipeline under scripts/perios/")
    ap.add_argument("--skip-viz", action="store_true", help="Skip the GIF visualization stage")
    ap.add_argument("--spacing-um", type=float, default=4.0,
                    help="Voxel size in micrometres (used for the mm conversions)")
    args = ap.parse_args()

    case = re.sub(r"[^A-Za-z0-9._-]+", "_", args.case).strip("_") or "case"
    if case != args.case:
        log(f"[prepare] case name sanitized: {args.case!r} -> {case!r}")
    out = os.path.abspath(args.out)
    os.makedirs(out, exist_ok=True)

    folder, script = resolve_pipeline(args.pipeline)
    log(f"=== morphometry: case={case} pipeline={args.pipeline} "
        f"spacing={args.spacing_um} um (CPU) ===")

    # 1) stage the inputs the way the pipeline expects
    t0 = time.time()
    stage_inputs(args.mask, args.image, out, case)
    prepare_sec = time.time() - t0

    # 2) run the vendored pipeline. -u so its stage output interleaves with ours in
    #    the worker's log file (both are writing to the same non-tty handle).
    cmd = [sys.executable, "-u", script, out]
    if args.skip_viz:
        cmd.append("--skip-viz")
    log(f"=== running {args.pipeline} ===")
    log("$ " + " ".join(cmd))
    sys.stdout.flush()
    env = dict(os.environ)
    env.setdefault("PYTHONIOENCODING", "utf-8")
    t1 = time.time()
    # Tee rather than inherit stdout: run_pipeline.py catches every per-stage
    # exception, prints "[FAIL] <stage>", and still exits 0 — so its exit code is
    # not evidence of anything. Its own stage-status lines are, and we can only
    # read them if we capture them. Still echoed line by line so the worker's log
    # looks exactly as before.
    stage_failures: list[str] = []
    try:
        proc = subprocess.Popen(cmd, cwd=folder, env=env, stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT, text=True,
                                encoding="utf-8", errors="replace", bufsize=1)
    except OSError as e:
        die(f"could not start the {args.pipeline} pipeline: {e}")
        return
    assert proc.stdout is not None
    for line in proc.stdout:
        sys.stdout.write(line)
        # Flush per line. The child used to write straight to the inherited log
        # handle, so its progress appeared live; routing it through this process
        # means our own block buffering would otherwise hold a long-running
        # stage's output back for minutes, and the UI tails this log.
        sys.stdout.flush()
        stripped = line.strip()
        if any(mark in stripped for mark in PIPELINE_FAILURE_MARKERS):
            stage_failures.append(stripped)
    rc = proc.wait()
    pipeline_sec = time.time() - t1
    sys.stdout.flush()
    log(f"=== {args.pipeline} exited with code {rc} in {pipeline_sec:.0f}s ===")
    if rc != 0:
        die(f"{args.pipeline}/run_pipeline.py failed with exit code {rc} — "
            f"see the stage output above")

    # 3) consolidate. run_pipeline.py swallows per-stage exceptions and still exits
    #    0, so a zero exit code proves nothing: validate that metrics actually exist.
    t2 = time.time()
    mdir = os.path.join(out, "metrics")
    if not os.path.isdir(mdir):
        die(f"the pipeline produced no metrics/ directory in {out} — every stage failed")
    timings = {"prepare": round(prepare_sec, 1), "pipeline": round(pipeline_sec, 1)}
    metrics, raw, annotated, xlsx_path = consolidate(
        out, case, args.spacing_um, args.pipeline, args.skip_viz, timings)
    if not raw:
        die(f"the pipeline wrote no readable metrics JSON into {mdir} — every stage failed")

    # A stage that crashed is a FAILED measurement, not a partial one. Reporting
    # "succeeded" with a metric silently missing is the worst outcome available:
    # the run looks fine, the gap is invisible, and GET /stats quietly averages
    # over a smaller n as though the sample had simply not been measured.
    if stage_failures:
        die(f"the pipeline reported a failed stage for case {case!r} despite exiting 0:\n  "
            + "\n  ".join(stage_failures)
            + f"\n(metrics files present: {sorted(os.listdir(mdir))})")

    socket_v = metrics.get("socket_volume_voxels")
    bone_v = metrics.get("bone_length_voxels")

    if socket_v is None:
        die(f"stage 4 produced no socket metrics for case {case!r} in {mdir} — "
            f"the pipeline ran but produced nothing measurable "
            f"(files present: {sorted(os.listdir(mdir))})")

    if socket_v == 0:
        # A specimen with no detectable socket is a real scientific result. But
        # utils.compute_3d_convex_hull swallows every exception — including
        # MemoryError — and a failed hull ALSO yields a socket volume of zero.
        # The two are told apart by the hull itself: a working hull always
        # encloses more than the shrunk mask, so a hull that does not is a
        # crashed hull, not a socketless bone.
        hull3d = metrics.get("hull3d_volume_voxels")
        shrunk = metrics.get("phalanx_volume_voxels")
        if hull3d is None or (shrunk is not None and hull3d <= shrunk):
            die(f"stage 4 reported zero socket volume for case {case!r}, but its 3D "
                f"convex hull ({hull3d}) does not enclose the shrink-wrapped mask "
                f"({shrunk}) — the hull computation failed rather than finding no "
                f"socket. utils.compute_3d_convex_hull swallows the real exception; "
                f"see the stage output above.")
        metrics["no_socket_detected"] = True
        log(f"[note] no socket detected for {case!r} — a valid result, but the socket "
            f"and bone-length metrics are undefined for this sample.")
    elif bone_v is None:
        # A real socket was found, so stages 5/6 had everything they needed and
        # should have produced a length. Their absence means they died.
        die(f"stage 4 found a socket ({socket_v:g} voxels) for case {case!r} but no "
            f"bone-length metrics were written — stages 5/6 failed silently "
            f"(files present: {sorted(os.listdir(mdir))})")

    # Advisory plausibility checks — recorded, never fatal. The reviewer decides.
    #
    # `qc_warnings` and `qc_checked` are ALWAYS written, including when nothing
    # tripped. Writing the key only on failure would make "checked, all clear"
    # and "never checked" byte-identical to a reader, and a UI cannot then tell
    # an all-clear from an unknown — it would show a clean bill of health over
    # numbers nothing ever looked at. That is the one outcome this whole feature
    # exists to prevent, so the distinction is recorded explicitly.
    try:
        sys.path.insert(0, os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src"))
        from microct_lab.morphqc import evaluate
        qc = evaluate(metrics)
        metrics["qc_warnings"] = qc
        metrics["qc_checked"] = True
        if qc:
            log(f"[qc] {len(qc)} plausibility warning(s):")
            for w in qc:
                log(f"[qc]   {w['severity']:6s} {w['code']}: {w['message']}")
        else:
            log("[qc] no plausibility warnings — values lie inside the reference range")
    except Exception as e:  # noqa: BLE001 — advisory only, never fail the run for it
        metrics["qc_checked"] = False
        metrics["qc_error"] = f"{type(e).__name__}: {e}"
        log(f"[qc] plausibility checks UNAVAILABLE: {type(e).__name__}: {e}")
        log("[qc] these metrics are unchecked — treat them as unverified.")
    parse_sec = time.time() - t2
    timings["parse"] = round(parse_sec, 1)
    timings["total"] = round(prepare_sec + pipeline_sec + parse_sec, 1)
    metrics.update({f"{k}_seconds": v for k, v in timings.items()})

    result = {
        "case": case,
        "pipeline_version": args.pipeline,
        "spacing_um": args.spacing_um,
        "skip_viz": bool(args.skip_viz),
        "output_dir": out,
        "annotated_nii": annotated,
        "annotated_labels": {str(k): v for k, v in ANNOTATED_LABELS.items()},
        "xlsx_path": xlsx_path,
        "metrics": metrics,
        "raw": raw,
        "environment": capture_env(args.pipeline, timings),
    }
    result_path = os.path.join(out, f"{case}_measurement.json")
    with open(result_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, default=str)
    log(f"[metrics] {len(metrics)} keys -> {result_path}")

    def fmt(v, unit="", nd=4):
        return f"{v:,.{nd}f}{unit}" if isinstance(v, (int, float)) else "n/a"

    log(f"[result] {case}: socket = {fmt(metrics.get('socket_volume_voxels'), '', 0)} voxels "
        f"= {fmt(metrics.get('socket_volume_mm3'), ' mm^3')}")
    log(f"[result] {case}: socket radius = {fmt(metrics.get('socket_radius_voxels'), '', 1)} voxels "
        f"= {fmt(metrics.get('socket_radius_mm'), ' mm')}")
    log(f"[result] {case}: phalanx = {fmt(metrics.get('phalanx_volume_voxels'), '', 0)} voxels "
        f"= {fmt(metrics.get('phalanx_volume_mm3'), ' mm^3')}")
    log(f"[result] {case}: bone length = {fmt(metrics.get('bone_length_voxels'), '', 1)} voxels "
        f"= {fmt(metrics.get('bone_length_mm'), ' mm')}")
    log(f"[result] {case}: euclidean = {fmt(metrics.get('euclidean_distance_voxels'), '', 1)} voxels "
        f"= {fmt(metrics.get('euclidean_distance_mm'), ' mm')}")
    if annotated:
        log(f"[result] annotated 7-class volume -> {annotated}")
    else:
        log("[result] WARNING: no annotated volume was produced (stage 5/6 failed)")
    log("DONE")


if __name__ == "__main__":
    main()
